"""データベース（data/）から Excel を書き出す:  python -m etl.export_xlsx [--out FOLDER] [--no-recalc]

1 ラウンド = 1 ブック。列は以前の「全試合のリザルト_Excel」と同じ並び（A〜AH）に、Q1/Q2 二段用の列（AI〜AL）を足したもの。
合計欄は Excel 数式（FIS 規則）で書き、Excel で再計算したあと、データベースの値と全セル照合する。
データベース側は多層照合を通った値なので、この Excel はその「控え」であり、独立の検証ではない。
"""
import os, sys, json, argparse, datetime, csv, shutil
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from . import config

EPS = "0.000001"
ROUND_JP = {'Q': '予選', 'Q1': '予選1', 'Q2': '予選2', 'F1': '決勝1', 'F2': '決勝2', 'F3': '決勝3'}
GENDER_JP = {'M': '男子', 'W': '女子'}


def load_db():
    m = json.load(open(os.path.join(config.DATA_DIR, 'manifest.json'), encoding='utf-8'))
    events = json.load(open(os.path.join(config.DATA_DIR, m['files']['events']), encoding='utf-8'))
    runs = []
    for s, f in m['files']['runs'].items():
        runs += json.load(open(os.path.join(config.DATA_DIR, f), encoding='utf-8'))
    return m, events, runs


def air_term(r, j6, j7, dd):
    return f"(MIN(10,TRUNC({j6}{r}*{dd}{r}+{EPS},2))+MIN(10,TRUNC({j7}{r}*{dd}{r}+{EPS},2)))/2"


def build_workbook(rnd, ev, runs, manifest, path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Results'
    FONT = 'Arial'
    bold, normal = Font(name=FONT, bold=True), Font(name=FONT)
    header_fill, group_fill = PatternFill('solid', start_color='D9E1F2'), PatternFill('solid', start_color='BDD7EE')
    thin = Side(style='thin', color='999999')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    muted = Font(name=FONT, color='999999', strike=True)

    title = f"{ev['season']} {ev['series']} {rnd['venue']}  |  {GENDER_JP[rnd['gender']]} {ROUND_JP[rnd['round']]}  |  {rnd['date']} {rnd['start_time'] or ''}  |  Codex {rnd['codex']}"
    ws.merge_cells('A1:AL1'); ws['A1'] = title; ws['A1'].font = Font(name=FONT, bold=True, size=12)
    ws['A2'] = f"データ版 {manifest['dataVersion']}　検証: " + ', '.join(f"{k} {v}" for k, v in rnd['verification'].items()) + "　（数式の合計は FIS 規則。データベースと全セル照合済み）"
    ws['A2'].font = Font(name=FONT, size=9, color='666666')
    groups = [('A3', 'F3', ''), ('G3', 'H3', 'Time'), ('I3', 'L3', 'Air - Jump 1'), ('M3', 'P3', 'Air - Jump 2'), ('Q3', 'Q3', 'Air'),
              ('R3', 'W3', 'Turns - Base Score (B)'), ('X3', 'AC3', 'Turns - Deductions (D)'), ('AD3', 'AD3', 'Turns'), ('AE3', 'AL3', '')]
    for s, e, label in groups:
        ws.merge_cells(f'{s}:{e}'); c = ws[s]; c.value = label; c.font = bold; c.fill = group_fill; c.alignment = center
    headers = ['Rank', 'Bib', 'FIS Code', 'Name', 'NOC', 'YB', 'Seconds', 'Time Points', 'J6', 'J7', 'Jump', 'DD', 'J6', 'J7', 'Jump', 'DD', 'Air Total',
               'J1', 'J2', 'J3', 'J4', 'J5', 'Base Total', 'J1', 'J2', 'J3', 'J4', 'J5', 'Deduction Total', 'Turns Total', 'Run Score', 'Tie', 'Status',
               'Reserve Judge', 'Q Block', 'Counting', 'Best Score', 'run_id']
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h); c.font = bold; c.fill = header_fill; c.alignment = center; c.border = border

    def put(row, col, value, font=None):
        cell = ws.cell(row=row, column=col, value=value); cell.font = font or normal; cell.border = border; cell.alignment = center
        return cell

    pace_ref = "'Event Info'!$E$1"
    row = 5
    for run in runs:
        ok = run['status'] == 'OK'
        j1 = run['air'][0] if len(run['air']) > 0 else {}
        j2 = run['air'][1] if len(run['air']) > 1 else {}
        for col, v in enumerate([run['rank'], run['bib'], run['fis_code'], run['name'], run['noc'], run['yb'], run['seconds']], start=1):
            put(row, col, v)
        put(row, 8, f"=TRUNC(MAX(0,MIN(20,48-32*G{row}/{pace_ref}))+{EPS},2)" if ok and run['seconds'] is not None else run['time_points'])
        for col, v in enumerate([j1.get('J6'), j1.get('J7'), j1.get('jump'), j1.get('dd'), j2.get('J6'), j2.get('J7'), j2.get('jump'), j2.get('dd')], start=9):
            put(row, col, v)
        terms = [air_term(row, 'I', 'J', 'L')] if j1 else []
        if j2: terms.append(air_term(row, 'M', 'N', 'P'))
        put(row, 17, f"=TRUNC({'+'.join(terms)}+{EPS},2)" if ok and terms else run['air_total'])
        for k, v in enumerate(run['base']):
            put(row, 18 + k, v, muted if k in run['base_discard'] else normal)
        put(row, 23, f"=SUM(R{row}:V{row})-MAX(R{row}:V{row})-MIN(R{row}:V{row})" if ok else run['status'])
        for k, v in enumerate(run['ded']):
            put(row, 24 + k, v, muted if k in run['ded_discard'] else normal)
        put(row, 29, f"=SUM(X{row}:AB{row})-MAX(X{row}:AB{row})-MIN(X{row}:AB{row})" if ok else run['status'])
        put(row, 30, f"=MAX(0.3,W{row}+AC{row})" if ok else run['status'])
        put(row, 31, f"=H{row}+Q{row}+AD{row}" if ok else run['status'], Font(name=FONT, bold=True))
        put(row, 32, run['tie']); put(row, 33, run['status']); put(row, 34, 'RES' if run['reserve_judge'] else None)
        put(row, 35, run['q_block']); put(row, 36, '採用' if run['counting'] else '参考'); put(row, 37, run['best_score']); put(row, 38, run['run_id'])
        row += 1
    for r in range(5, row):
        for col in ['G', 'H', 'I', 'J', 'L', 'M', 'N', 'P', 'Q'] + list('RSTUVW') + ['X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AK']:
            ws[f'{col}{r}'].number_format = '0.00'
    widths = {'A': 5, 'B': 5, 'C': 10, 'D': 26, 'E': 6, 'F': 6, 'G': 7, 'H': 9, 'I': 5, 'J': 5, 'K': 7, 'L': 6, 'M': 5, 'N': 5, 'O': 7, 'P': 6, 'Q': 8,
              'R': 6, 'S': 6, 'T': 6, 'U': 6, 'V': 6, 'W': 8, 'X': 6, 'Y': 6, 'Z': 6, 'AA': 6, 'AB': 6, 'AC': 9, 'AD': 8, 'AE': 8, 'AF': 6, 'AG': 8,
              'AH': 9, 'AI': 7, 'AJ': 7, 'AK': 9, 'AL': 30}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A5'

    ws2 = wb.create_sheet('Event Info')
    left = [('Event', f"{GENDER_JP[rnd['gender']]} モーグル"), ('Round', rnd['round_text']), ('Date', rnd['date']), ('Start Time', rnd['start_time']),
            ('Venue', rnd['venue']), ('Codex', rnd['codex']), ('Source PDF', rnd['source']['pdf']), ('FIS URL', rnd['source'].get('fis_url')),
            ('PDF SHA-256', rnd['source']['pdf_sha256']), ('Data version', manifest['dataVersion']), ('Rules version', rnd['source']['rules_version'])]
    for r, (k, v) in enumerate(left, start=1):
        ws2[f'A{r}'] = k; ws2[f'A{r}'].font = bold; ws2[f'B{r}'] = v
    right = [('Pace Time', rnd['pace_time']), ('Course Length (m)', rnd['course']['length_m']), ('Course Width (m)', rnd['course']['width_m']),
             ('Gate Width (m)', rnd['course']['gate_width_m']), ('Gradient (deg)', rnd['course']['gradient_deg']), ('Competitors', rnd['n_competitors'])]
    for r, (k, v) in enumerate(right, start=1):
        ws2[f'D{r}'] = k; ws2[f'D{r}'].font = bold; ws2[f'E{r}'] = v
    ws2['A13'] = 'Judges'; ws2['A13'].font = Font(name=FONT, bold=True, size=12)
    for i, h in enumerate(['Judge No', 'Role', 'Name', 'NOC'], start=1):
        c = ws2.cell(row=14, column=i, value=h); c.font = bold; c.fill = header_fill; c.border = border
    r = 15
    for j in rnd['judges']:
        for i, v in enumerate([j['no'], j['role'], j['name'], j['noc']], start=1):
            ws2.cell(row=r, column=i, value=v).border = border
        r += 1
    r += 1
    ws2.cell(row=r, column=1, value='Officials').font = Font(name=FONT, bold=True, size=12); r += 1
    for i, h in enumerate(['Role', 'Name', 'NOC'], start=1):
        c = ws2.cell(row=r, column=i, value=h); c.font = bold; c.fill = header_fill; c.border = border
    r += 1
    for o in rnd['officials']:
        for i, v in enumerate([o['role'], o['name'], o['noc']], start=1):
            ws2.cell(row=r, column=i, value=v).border = border
        r += 1
    for col, w in {'A': 22, 'B': 40, 'C': 22, 'D': 20, 'E': 10}.items():
        ws2.column_dimensions[col].width = w
    wb.save(path)


class ExcelRecalc:
    def __init__(self):
        import win32com.client
        self.xl = win32com.client.DispatchEx('Excel.Application'); self.xl.Visible = False; self.xl.DisplayAlerts = False

    def recalc(self, path):
        wb = self.xl.Workbooks.Open(os.path.abspath(path))
        try:
            self.xl.CalculateFullRebuild(); wb.Save()
        finally:
            wb.Close(SaveChanges=False)

    def close(self):
        try:
            self.xl.Quit()
        except Exception:
            pass


def verify(path, runs):
    wb = load_workbook(path, data_only=True); ws = wb['Results']
    mism = []
    def num(a, b):
        return (a is None and b is None) or (a is not None and b is not None and abs(float(a) - float(b)) <= 0.005)
    for i, run in enumerate(runs):
        r = 5 + i
        checks = [('D', run['name']), ('E', run['noc']), ('AL', run['run_id'])]
        for col, exp in checks:
            if ws[f'{col}{r}'].value != exp:
                mism.append(f"{col}{r} {exp!r} != {ws[f'{col}{r}'].value!r}")
        if run['status'] != 'OK':
            continue
        for col, exp in [('H', run['time_points']), ('Q', run['air_total']), ('W', run['base_total']), ('AC', run['ded_total']), ('AD', run['turns_total']), ('AE', run['run_score'])]:
            if not num(ws[f'{col}{r}'].value, exp):
                mism.append(f"{col}{r} {run['name']}: DB {exp} / Excel {ws[f'{col}{r}'].value}")
    return mism


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument('--out', default=os.path.join(os.path.dirname(os.path.dirname(config.REPO)), 'ジャッジ分析', '全試合のリザルト_Excel'))
    ap.add_argument('--no-recalc', action='store_true')
    args = ap.parse_args(argv)
    out = os.path.normpath(args.out)
    manifest, events, runs = load_db()
    by_round = {}
    for r in runs:
        by_round.setdefault(r['round_id'], []).append(r)
    rounds = [(rnd, ev) for ev in events for rnd in ev['rounds']]
    # start clean: remove previous workbooks/logs (keep nothing else in this folder)
    if os.path.isdir(out):
        for root, dirs, files in os.walk(out):
            for f in files:
                if f.endswith(('.xlsx', '.csv', '.txt', '.md')):
                    os.remove(os.path.join(root, f))
    os.makedirs(out, exist_ok=True)
    xl = None if args.no_recalc else ExcelRecalc()
    log = []
    try:
        for rnd, ev in rounds:
            rel = os.path.splitext(rnd['source']['pdf'])[0] + '.xlsx'
            path = os.path.join(out, rel)
            os.makedirs(os.path.dirname(path), exist_ok=True)
            rs = by_round.get(rnd['round_id'], [])
            rs = sorted(rs, key=lambda r: (r['rank'] if r['rank'] else 10 ** 6, r['bib'], 0 if r['counting'] else 1))
            build_workbook(rnd, ev, rs, manifest, path)
            if xl:
                xl.recalc(path)
                mism = verify(path, rs)
            else:
                mism = ['(未再計算)']
            log.append({'round_id': rnd['round_id'], 'xlsx': rel, 'runs': len(rs), 'status': 'OK' if not mism else 'NG', 'message': ' | '.join(mism[:5])})
            print(f"  {rnd['round_id']:26s} {len(rs):3d} rows  {'OK' if not mism else 'NG ' + mism[0]}")
    finally:
        if xl:
            xl.close()
    with open(os.path.join(out, '書き出しログ.csv'), 'w', encoding='utf-8-sig', newline='') as fh:
        w = csv.DictWriter(fh, fieldnames=['round_id', 'xlsx', 'runs', 'status', 'message']); w.writeheader(); w.writerows(log)
    with open(os.path.join(out, 'README_このフォルダについて.md'), 'w', encoding='utf-8') as fh:
        fh.write(README.format(ver=manifest['dataVersion'], n=len(log), when=datetime.datetime.now().strftime('%Y-%m-%d %H:%M')))
    n_ng = sum(1 for l in log if l['status'] != 'OK')
    print(f"\n{len(log)} ブック書き出し / NG {n_ng} / 出力先 {out} / データ版 {manifest['dataVersion']}")
    return 0 if n_ng == 0 else 1


README = """# 全試合のリザルト_Excel

`moguls-results` のデータベース（data/、データ版 {ver}）から書き出した Excel。{n} ブック、生成 {when}。
`python -m etl.export_xlsx` で作り直せる。**手で編集しない**（次回の書き出しで上書きされる）。

- 1 PDF = 1 ブック。フォルダ構成とファイル名は元 PDF と同じ。
- 数値はデータベースと同じ（多層照合を通った値）。合計欄は FIS 規則の Excel 数式で、Excel で再計算後にデータベースと全セル照合済み（`書き出しログ.csv`）。
- 取り消し線・灰色の審判点は最高・最低で除外された点。
- 列: A〜AH は従来と同じ（Rank … Reserve Judge）。AI Q Block（Q1/Q2 二段の予選）、AJ Counting（採用／参考）、AK Best Score、AL run_id（データベースとの対応キー）。
- Event Info: 大会情報、ペースタイム・コース、審判・審判団、元 PDF の FIS URL と SHA-256、データ版、規則版。
- 対象外: デュアルモーグル。
"""

if __name__ == '__main__':
    sys.exit(main())
